"""Regression tests for export utility helpers."""

from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pandas as pd
import pytest
from openpyxl import Workbook

import OpenPinch.utils.export as export_mod
from OpenPinch.classes import ProblemTable, Zone
from OpenPinch.lib.enums import ProblemTableLabel as PT
from OpenPinch.lib.schemas.common import PeriodValueWithUnit
from OpenPinch.lib.schemas.targets import DirectIntegrationTarget
from OpenPinch.utils.export import (
    _autosize_columns,
    _safe_name,
    build_summary_dataframe,
    export_target_summary_to_excel_with_units,
)

# --------------------------------------------------------------------------------------
# Fixtures & tiny helpers
# --------------------------------------------------------------------------------------


class VU:
    """Simple value-with-units stub."""

    def __init__(self, value, unit):
        self.value = value
        self.unit = unit


def _make_target(
    name="Base",
    period_idx=None,
    period_id=None,
    pinch_cold=VU(85.0, "degC"),
    pinch_hot=VU(125.0, "degC"),
    Qh=VU(100.0, "kW"),
    Qc=VU(80.0, "kW"),
    Qr=VU(20.0, "kW"),
    degree_of_integration=VU(70.0, "%"),
    utility_cost=VU(1234.0, "$/h"),
    area=VU(456.0, "m^2"),
    work_target=VU(10.0, "kW"),
    turbine_eff_target=VU(80.0, "%"),
    exergy_sources=VU(200.0, "kW"),
    exergy_sinks=VU(180.0, "kW"),
    exergy_req_min=VU(15.0, "kW"),
    exergy_des_min=VU(12.0, "kW"),
    num_units=7,
    capital_cost=VU(111.0, "$"),
    total_cost=VU(222.0, "$/y"),
    ETE=VU(93.0, "%"),
    hot_utils=None,
    cold_utils=None,
):
    """Builds a single TargetResults-like stub compatible with export.py access pattern."""
    if hot_utils is None:
        hot_utils = [SimpleNamespace(name="HP Steam", heat_flow=VU(75.0, "kW"))]
    if cold_utils is None:
        cold_utils = [SimpleNamespace(name="Cooling Water", heat_flow=VU(60.0, "kW"))]

    return SimpleNamespace(
        name=name,
        period_idx=period_idx,
        period_id=period_id,
        pinch_temp=SimpleNamespace(cold_temp=pinch_cold, hot_temp=pinch_hot),
        Qh=Qh,
        Qc=Qc,
        Qr=Qr,
        degree_of_integration=degree_of_integration,
        hot_utilities=hot_utils,
        cold_utilities=cold_utils,
        utility_cost=utility_cost,
        area=area,
        work_target=work_target,
        turbine_efficiency_target=turbine_eff_target,
        exergy_sources=exergy_sources,
        exergy_sinks=exergy_sinks,
        exergy_req_min=exergy_req_min,
        exergy_des_min=exergy_des_min,
        num_units=num_units,
        capital_cost=capital_cost,
        total_cost=total_cost,
        ETE=ETE,
    )


def _make_problem_table(values):
    """Small ProblemTable with a few populated numeric columns."""
    return ProblemTable(
        {
            PT.T: values,
            PT.H_HOT: [v * 1.1 for v in values],
            PT.H_COLD: [v * 2.2 for v in values],
        }
    )


# --------------------------------------------------------------------------------------
# export_target_summary_to_excel_with_units
# --------------------------------------------------------------------------------------


def test_export_writes_expected_excel(tmp_path: Path, monkeypatch):
    """
    End-to-end: ensure the Excel file is written with the right filename pattern
    and the Summary sheet contains expected columns/values (values + units).
    """

    # Freeze timestamp used in filename
    class _FixedDT:
        @staticmethod
        def now():
            return datetime(2025, 1, 2, 3, 4, 5)

        @staticmethod
        def strptime(*a, **kw):
            return datetime.strptime(*a, **kw)

    # Patch the datetime reference inside the export module

    monkeypatch.setattr(export_mod, "datetime", _FixedDT, raising=True)

    # Project name exercises _safe_name (forbidden chars + spaces)
    target_response = SimpleNamespace(
        name="Proj: Foo/Bar",  # becomes Proj__Foo_Bar
        targets=[
            _make_target(name="Base"),
            _make_target(
                name="Alt A",
                hot_utils=[SimpleNamespace(name="MP Steam", heat_flow=VU(55.0, "kW"))],
                cold_utils=[
                    SimpleNamespace(name="Chilled Water", heat_flow=VU(30.0, "kW"))
                ],
            ),
        ],
    )

    out = export_target_summary_to_excel_with_units(
        target_response, master_zone=None, out_dir=tmp_path
    )
    out_path = Path(out)
    assert out_path.exists()

    # Filename pattern check
    assert out_path.name == "Proj__Foo_Bar_20250102_030405.xlsx"

    # Verify Summary sheet contents round-trip
    df = pd.read_excel(out_path, sheet_name="Summary")
    # Expect at least our known columns
    expected_cols = {
        "Target",
        "Cold Pinch (value)",
        "Cold Pinch (unit)",
        "Hot Pinch (value)",
        "Hot Pinch (unit)",
        "Qh (value)",
        "Qh (unit)",
        "Qc (value)",
        "Qc (unit)",
        "Qr (value)",
        "Qr (unit)",
        "Degree of Integration (value)",
        "Degree of Integration (unit)",
        "Utility Cost (value)",
        "Utility Cost (unit)",
        "Area (value)",
        "Area (unit)",
        "Num Units",
        "Capital Cost (value)",
        "Capital Cost (unit)",
        "Total Cost (value)",
        "Total Cost (unit)",
        "Work Target (value)",
        "Work Target (unit)",
        "Turbine Eff Target (value)",
        "Turbine Eff Target (unit)",
        "ETE (value)",
        "ETE (unit)",
        "Exergy Sources (value)",
        "Exergy Sources (unit)",
        "Exergy Sinks (value)",
        "Exergy Sinks (unit)",
        "Exergy Req Min (value)",
        "Exergy Req Min (unit)",
        "Exergy Des Min (value)",
        "Exergy Des Min (unit)",
        "HPR Cycle",
        "HPR Success",
        "HPR Utility Total (value)",
        "HPR Utility Total (unit)",
        "HPR Work (value)",
        "HPR Work (unit)",
        "HPR External Utility (value)",
        "HPR External Utility (unit)",
        "HPR Ambient Hot (value)",
        "HPR Ambient Hot (unit)",
        "HPR Ambient Cold (value)",
        "HPR Ambient Cold (unit)",
        "HPR COP (value)",
        "HPR COP (unit)",
        "HPR Eta HE (value)",
        "HPR Eta HE (unit)",
        "HPR Operating Cost (value)",
        "HPR Operating Cost (unit)",
        "HPR Capital Cost (value)",
        "HPR Capital Cost (unit)",
        "HPR Annualized Capital Cost (value)",
        "HPR Annualized Capital Cost (unit)",
        "HPR Total Annualized Cost (value)",
        "HPR Total Annualized Cost (unit)",
        "HPR Compressor Capital Cost (value)",
        "HPR Compressor Capital Cost (unit)",
        "HPR Heat Exchanger Capital Cost (value)",
        "HPR Heat Exchanger Capital Cost (unit)",
        # Utility columns come from names:
        "HP Steam (value)",
        "HP Steam (unit)",
        "Cooling Water (value)",
        "Cooling Water (unit)",
        "MP Steam (value)",
        "MP Steam (unit)",
        "Chilled Water (value)",
        "Chilled Water (unit)",
    }
    assert expected_cols.issubset(set(df.columns))

    # Row for "Base"
    base = df.loc[df["Target"] == "Base"].iloc[0]
    assert pytest.approx(base["Cold Pinch (value)"]) == 85.0
    assert base["Cold Pinch (unit)"] == "degC"
    assert pytest.approx(base["Qh (value)"]) == 100.0
    assert base["Qh (unit)"] == "kW"
    assert pytest.approx(base["HP Steam (value)"]) == 75.0
    assert base["HP Steam (unit)"] == "kW"
    assert pytest.approx(base["Cooling Water (value)"]) == 60.0
    assert base["Cooling Water (unit)"] == "kW"
    assert int(base["Num Units"]) == 7
    assert pytest.approx(base["Capital Cost (value)"]) == 111.0
    assert base["Capital Cost (unit)"] == "$"
    assert pytest.approx(base["ETE (value)"]) == 93.0
    assert base["ETE (unit)"] == "%"

    # Row for "Alt A"
    alt = df.loc[df["Target"] == "Alt A"].iloc[0]
    assert pytest.approx(alt["MP Steam (value)"]) == 55.0
    assert alt["MP Steam (unit)"] == "kW"
    assert pytest.approx(alt["Chilled Water (value)"]) == 30.0
    assert alt["Chilled Water (unit)"] == "kW"


def test_export_writes_problem_tables_for_all_zones(tmp_path: Path):
    master_zone = Zone("Plant")
    master_target = DirectIntegrationTarget(
        zone_name="Master",
        type="DI",
        pt=_make_problem_table([10.0]),
        pt_real=_make_problem_table([30.0]),
        hot_utility_target=0.0,
        cold_utility_target=0.0,
        heat_recovery_target=0.0,
    )
    master_target.pt = _make_problem_table([10.0, 20.0])
    master_target.pt_real = _make_problem_table([30.0])
    master_zone.targets["DI"] = master_target

    sub_zone = Zone("Sub/Zone", parent_zone=master_zone)
    master_zone.add_zone(sub_zone, sub=True)
    sub_target = DirectIntegrationTarget(
        zone_name="Alt:Target",
        type="DI",
        parent_zone=sub_zone,
        pt=_make_problem_table([40.0]),
        pt_real=_make_problem_table([50.0]),
        hot_utility_target=0.0,
        cold_utility_target=0.0,
        heat_recovery_target=0.0,
    )
    sub_target.pt = _make_problem_table([40.0])
    sub_target.pt_real = _make_problem_table([50.0])
    sub_zone.targets["DI"] = sub_target

    target_response = SimpleNamespace(name="Project", targets=[])
    out = export_target_summary_to_excel_with_units(
        target_response, master_zone=master_zone, out_dir=tmp_path
    )
    xls = pd.ExcelFile(out)

    master_shifted = "Plant - DI (Shifted)"
    master_real = "Plant - DI (Real)"
    sub_shifted = "Sub_Zone - DI (Shifted)"
    sub_real = "Sub_Zone - DI (Real)"

    assert master_shifted in xls.sheet_names
    assert master_real in xls.sheet_names
    assert sub_shifted in xls.sheet_names
    assert sub_real in xls.sheet_names

    wb = openpyxl.load_workbook(out, data_only=True)
    assert wb[master_shifted]["A1"].value == "Master/DI"
    assert wb[sub_shifted]["A1"].value == "Alt:Target/DI"

    master_df = pd.read_excel(
        out,
        sheet_name=master_shifted,
        usecols=lambda c: not str(c).startswith("Unnamed"),
        header=2,
    )
    assert pytest.approx(master_df.iloc[0][PT.T.value]) == 10.0
    assert pytest.approx(master_df.iloc[0][PT.H_HOT.value]) == 11.0

    sub_df = pd.read_excel(
        out,
        sheet_name=sub_shifted,
        usecols=lambda c: not str(c).startswith("Unnamed"),
        header=2,
    )
    assert pytest.approx(sub_df.iloc[0][PT.T.value]) == 40.0


def test_build_summary_dataframe_resolves_period_values_using_period_idx():
    target = _make_target(
        name="Peak",
        period_idx=1,
        period_id="peak",
        pinch_cold=PeriodValueWithUnit(values=[80.0, 95.0], unit="degC"),
        pinch_hot=PeriodValueWithUnit(values=[120.0, 135.0], unit="degC"),
        Qh=PeriodValueWithUnit(values=[100.0, 140.0], unit="kW"),
        Qc=PeriodValueWithUnit(values=[80.0, 60.0], unit="kW"),
        Qr=PeriodValueWithUnit(values=[20.0, 30.0], unit="kW"),
    )

    row = build_summary_dataframe([target]).iloc[0]

    assert row["Period ID"] == "peak"
    assert row["Qh (value)"] == pytest.approx(140.0)
    assert row["Qc (value)"] == pytest.approx(60.0)
    assert row["Qr (value)"] == pytest.approx(30.0)
    assert row["Cold Pinch (value)"] == pytest.approx(95.0)
    assert row["Hot Pinch (value)"] == pytest.approx(135.0)


def test_target_results_include_period_idx():
    target = DirectIntegrationTarget(
        zone_name="Plant",
        type="DI",
        pt=_make_problem_table([10.0]),
        pt_real=_make_problem_table([30.0]),
        hot_utility_target=10.0,
        cold_utility_target=5.0,
        heat_recovery_target=15.0,
        period_id="peak",
        period_idx=1,
    )

    results = target.to_target_results()

    assert results.period_idx == 1
    assert results.period_id == "peak"


# --------------------------------------------------------------------------------------
# _safe_name
# --------------------------------------------------------------------------------------


@pytest.mark.parametrize(
    "raw,expected",
    [
        (" Project  X  ", "Project_X"),
        ("data:run/01", "data_run_01"),
        ('a\\b|c*<d>?"e', "a_b_c_d_e"),
        ("", "Project"),
    ],
)
def test_safe_name_sanitization(raw, expected):
    assert _safe_name(raw) == expected


# --------------------------------------------------------------------------------------
# _autosize_columns
# --------------------------------------------------------------------------------------


def test_autosize_columns_sets_reasonable_widths(tmp_path: Path):
    # Prepare a small df with long header + long cell to exercise both code paths
    df = pd.DataFrame(
        {
            "A" * 50: ["x" * 10, "y" * 30],  # long header, varying cell content
            "B": ["short", "also short"],
        }
    )

    # Create a worksheet via openpyxl (used under the hood by pandas writer)

    wb = Workbook()
    ws = wb.active

    # Write header row so ws.cell(row=1, column=...) has a real cell
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=col)

    # Call autosize; should not raise and should cap widths at 40
    _autosize_columns(df, ws)

    # Check widths
    # First column width should be capped at 40 due to very long header
    first_col_letter = ws.cell(row=1, column=1).column_letter
    assert ws.column_dimensions[first_col_letter].width == 40

    # Second column should be at least len("B") + 2 = 3 and not exceed 40
    second_col_letter = ws.cell(row=1, column=2).column_letter
    w = ws.column_dimensions[second_col_letter].width
    assert 3 <= w <= 40


# ===== Merged from test_export_extra.py =====
"""Additional branch coverage tests for export helpers."""


def test_write_problem_tables_skips_empty_dataframes(monkeypatch, tmp_path: Path):
    zone = Zone("Plant")
    target = DirectIntegrationTarget(
        zone_name="Plant",
        type="DI",
        pt=_make_problem_table([1.0]),
        pt_real=_make_problem_table([2.0]),
        hot_utility_target=0.0,
        cold_utility_target=0.0,
        heat_recovery_target=0.0,
    )
    zone.targets["DI"] = target

    monkeypatch.setattr(
        export_mod,
        "problem_table_to_dataframe",
        lambda table, round_decimals=2: pd.DataFrame(),
    )

    out_path = tmp_path / "out.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        pd.DataFrame({"keep": [1]}).to_excel(writer, sheet_name="Keep", index=False)
        export_mod._write_problem_tables(zone, writer)
        assert set(writer.sheets.keys()) == {"Keep"}


def test_unique_sheet_name_suffix_and_exhaustion_paths():
    used = {"Sheet"}
    second = export_mod._unique_sheet_name("Sheet", used)
    assert second == "Sheet (2)"

    exhausted = {"Sheet"} | {f"Sheet ({i})" for i in range(2, 1000)}
    with pytest.raises(ValueError, match="Unable to allocate unique sheet name"):
        export_mod._unique_sheet_name("Sheet", exhausted)
