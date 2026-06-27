"""Regression tests for multi-period weighted summary outputs."""

from __future__ import annotations

from types import SimpleNamespace

import openpyxl
import pytest

from OpenPinch.classes._problem._multiperiod import (
    WEIGHTED_AVERAGE_PERIOD_ID,
    weighted_average_output,
)
from OpenPinch.classes.pinch_problem import PinchProblem
from OpenPinch.classes.value import Value
from OpenPinch.classes.zone import Zone
from OpenPinch.lib.enums import TT
from OpenPinch.lib.schemas.io import TargetOutput
from OpenPinch.lib.schemas.reporting import HeatUtility, PinchTemp, TargetResults


def _target(
    *,
    name: str = "Site/Direct Integration",
    period_id: str = "base",
    qh: float = 10.0,
    qc: float = 5.0,
    qr: float = 20.0,
    area: float | None = 100.0,
    hpr_success: bool | None = True,
    hot_utilities: list[HeatUtility] | None = None,
) -> TargetResults:
    return TargetResults(
        name=name,
        period_id=period_id,
        Qh=Value(qh, "kW"),
        Qc=Value(qc, "kW"),
        Qr=Value(qr, "kW"),
        utility_cost=Value(qh * 2.0, "$/h"),
        pinch_temp=PinchTemp(
            hot_temp=Value(120.0 + qh, "degC"),
            cold_temp=Value(90.0 + qh, "degC"),
        ),
        hot_utilities=(
            hot_utilities
            if hot_utilities is not None
            else [HeatUtility(name="Steam", heat_flow=Value(qh, "kW"))]
        ),
        cold_utilities=[HeatUtility(name="Cooling Water", heat_flow=Value(qc, "kW"))],
        work_target=Value(qh / 10.0, "kW"),
        process_component_work_target=Value(qh / 20.0, "kW"),
        turbine_efficiency_target=Value(40.0 + qh / 10.0, "%"),
        area=Value(area, "m^2") if area is not None else None,
        num_units=qh / 10.0,
        capital_cost=Value(qh * 100.0, "$"),
        total_cost=Value(qh * 10.0, "$/y"),
        exergy_sources=Value(qh * 0.8, "kW"),
        exergy_sinks=Value(qh * 0.6, "kW"),
        ETE=Value(50.0 + qh / 10.0, "%"),
        exergy_req_min=Value(qh * 0.2, "kW"),
        exergy_des_min=Value(qh * 0.1, "kW"),
        hpr_cycle="Carnot",
        hpr_utility_total=Value(qh * 1.1, "kW"),
        hpr_work=Value(qh * 0.1, "kW"),
        hpr_external_utility=Value(qh * 0.2, "kW"),
        hpr_ambient_hot=Value(qh * 0.3, "kW"),
        hpr_ambient_cold=Value(qh * 0.4, "kW"),
        hpr_cop=Value(3.0 + qh / 100.0, "-"),
        hpr_eta_he=Value(20.0 + qh / 10.0, "%"),
        hpr_operating_cost=Value(qh * 20.0, "$/y"),
        hpr_capital_cost=Value(qh * 30.0, "$"),
        hpr_annualized_capital_cost=Value(qh * 4.0, "$/y"),
        hpr_total_annualized_cost=Value(qh * 24.0, "$/y"),
        hpr_compressor_capital_cost=Value(qh * 12.0, "$"),
        hpr_heat_exchanger_capital_cost=Value(qh * 18.0, "$"),
        hpr_success=hpr_success,
    )


def test_weighted_average_output_aggregates_values_utilities_and_metadata():
    base = TargetOutput(
        name="Site",
        period_id="base",
        targets=[_target(period_id="base", qh=10.0, hpr_success=True)],
    )
    peak = TargetOutput(
        name="Site",
        period_id="peak",
        targets=[
            _target(
                period_id="peak",
                qh=30.0,
                hpr_success=False,
                hot_utilities=[
                    HeatUtility(name="Steam", heat_flow=Value(30.0, "kW")),
                    HeatUtility(name="Fuel", heat_flow=Value(8.0, "kW")),
                ],
            )
        ],
    )

    output = weighted_average_output([base, peak], [1.0, 3.0])
    target = output.targets[0]

    assert output.period_id == WEIGHTED_AVERAGE_PERIOD_ID
    assert target.period_id == WEIGHTED_AVERAGE_PERIOD_ID
    assert target.Qh.value == pytest.approx(25.0)
    assert target.Qc.value == pytest.approx(5.0)
    assert target.work_target.value == pytest.approx(2.5)
    assert target.process_component_work_target.value == pytest.approx(1.25)
    assert target.area.value == pytest.approx(100.0)
    assert target.capital_cost.value == pytest.approx(2500.0)
    assert target.exergy_sources.value == pytest.approx(20.0)
    assert target.hpr_work.value == pytest.approx(2.5)
    assert target.hpr_total_annualized_cost.value == pytest.approx(600.0)
    assert target.hpr_cycle == "Carnot"
    assert target.hpr_success is None
    assert {
        utility.name: utility.heat_flow.value for utility in target.hot_utilities
    } == {
        "Steam": pytest.approx(25.0),
        "Fuel": pytest.approx(6.0),
    }


def test_weighted_average_output_rejects_partially_missing_numeric_fields():
    base = TargetOutput(
        name="Site",
        period_id="base",
        targets=[_target(period_id="base", area=100.0)],
    )
    peak = TargetOutput(
        name="Site",
        period_id="peak",
        targets=[_target(period_id="peak", area=None)],
    )

    with pytest.raises(ValueError, match="partially missing field 'area'"):
        weighted_average_output([base, peak], [1.0, 1.0])


def test_weighted_summary_replays_last_named_target_accessor(monkeypatch):
    root = Zone("Site")
    root.set_period_context({"base": 0, "peak": 1}, [1.0, 3.0], 2)
    problem = PinchProblem()
    problem._master_zone = root
    calls = []

    def fake_execute_targeting(
        self,
        *,
        target_id,
        application_zone,
        options,
        include_subzones,
        **_kwargs,
    ):
        period_id = options["period_id"]
        calls.append((target_id, application_zone, include_subzones, period_id))
        qh = {"base": 10.0, "peak": 30.0}[period_id]
        self._results = TargetOutput(
            name="Site",
            period_id=period_id,
            targets=[_target(period_id=period_id, qh=qh)],
        )
        return SimpleNamespace(name="Site/Direct Integration")

    def fail_default_targeting(*_args, **_kwargs):
        raise AssertionError("default targeting should not run")

    monkeypatch.setattr(PinchProblem, "_execute_targeting", fake_execute_targeting)
    monkeypatch.setattr(
        PinchProblem,
        "_run_targeting_for_zone_and_subzones",
        fail_default_targeting,
    )

    problem.target.direct_heat_integration(
        zone_name="AreaA",
        include_subzones=True,
        period_id="peak",
    )
    frame = problem.summary_frame(periods="weighted_average", format="plain")

    assert calls == [
        (TT.DI.value, "AreaA", True, "peak"),
        (TT.DI.value, "AreaA", True, "base"),
        (TT.DI.value, "AreaA", True, "peak"),
    ]
    assert frame.iloc[0]["Period ID"] == WEIGHTED_AVERAGE_PERIOD_ID
    assert frame.iloc[0]["Hot Utility Target"] == pytest.approx(25.0)
    assert problem.results.period_id == "peak"


@pytest.mark.parametrize(
    "source",
    [
        "crude_preheat_train_multiperiod.json",
        "zonal_site_multiperiod.json",
    ],
)
def test_weighted_summary_matches_sample_period_weights(source):
    problem = PinchProblem(source=source, project_name=source)

    frame = problem.summary_frame(periods="weighted_average", format="plain")
    period_outputs = list(problem.target_all_periods().values())
    weights = list(problem.master_zone.weights)
    target_name = period_outputs[0].targets[0].name
    row = frame.loc[frame["Target"] == target_name].iloc[0]

    for column, attr_name in [
        ("Hot Utility Target", "Qh"),
        ("Cold Utility Target", "Qc"),
        ("Heat Recovery", "Qr"),
    ]:
        values = [
            float(getattr(output.targets[0], attr_name).value)
            for output in period_outputs
        ]
        expected = sum(weight * value for weight, value in zip(weights, values)) / sum(
            weights
        )
        assert row[column] == pytest.approx(expected)
    assert row["Period ID"] == WEIGHTED_AVERAGE_PERIOD_ID


def test_weighted_summary_surfaces_metrics_report_and_export(tmp_path):
    problem = PinchProblem(
        source="crude_preheat_train_multiperiod.json",
        project_name="weighted_summary_export",
    )

    metrics = problem.metrics(periods="weighted_average")
    report = problem.report(periods="weighted_average")
    export_path = problem.export_excel(tmp_path, periods="weighted_average")

    assert metrics
    assert {metric.period_id for metric in metrics} == {WEIGHTED_AVERAGE_PERIOD_ID}
    assert report.targets[0].period_id == WEIGHTED_AVERAGE_PERIOD_ID
    workbook = openpyxl.load_workbook(export_path, read_only=True)
    try:
        summary = workbook["Summary"]
        headers = [cell.value for cell in next(summary.iter_rows(max_row=1))]
        period_id_col = headers.index("Period ID") + 1
        assert summary.cell(row=2, column=period_id_col).value == (
            WEIGHTED_AVERAGE_PERIOD_ID
        )
    finally:
        workbook.close()
